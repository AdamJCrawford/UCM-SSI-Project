import csv
import config

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles.fills import PatternFill

import pandas as pd

import os
import globals
# This is the main piece of data collection. It reads in the file and creates a
# list of the all data that is required.


def fetch_data():
    with open(globals.STUDENT_DATA) as f:
        data = csv.reader(f)
        data = [row for row in data]
        data = data[2:]

        for i in range(len(data)):
            data[i][1] = data[i][1].split(",")
            length = len(data[i])
            # add comments
            data[i][2] = [data[i][2:length - 2]][0]
            data[i][3] = data[i][length - 2]
            data[i][4] = data[i][length - 1]
            data[i] = data[i][:5]
    return data


def get_project_data():
    project_data = []
    with open(globals.SKILLS_DATA) as f:
        data = csv.reader(f)
        for row in data:
            # Some times this random symbol will appear. Not sure what is is / why is appears. If it is there delete it.
            if "ï»¿" in row[0]:
                row[0] = row[0][-1]
            row = [int(row[0]), row[2].split(','), row[-1]]
            project_data.append(row)
        return project_data

# Reads the line that contains the line of project names and creates a list with
# the names.


def get_project_names():
    with open(globals.STUDENT_DATA) as f:
        next(csv.reader(f))
        data = next(csv.reader(f))[2:-2]

        for i in range(len(data)):
            # This replaces the question with "" (nothing) so the only thing left
            # is the name of the project
            data[i] = data[i].replace(
                "Please tell us how much you would like to work on each of these projects - 5 being MOST preferred and 1 being LEAST preferred. - ", "")

    return data


def fill_files():
    # This is the function that makes both of the output files
    with open("../../Fitness.csv", "w", newline='') as f:
        writer = csv.writer(f)

        headers = []
        offset = 0
        # Adds all of the col names to the excel file
        for i in range(config.LENGTH_PROJECT):
            if i != 0 and config.PROJECT_NAMES[i - 1] == config.PROJECT_NAMES[i]:
                offset += 1
            headers.append(config.TEAMS[i].identifier)
        headers.insert(0, "Email")
        writer.writerow(headers)

        for person in config.PEOPLE:
            # If the person is in a project add the '&' char to mark the cell for later
            # processing.
            row = [[str(person.ratings[i]) + '&'][0] if person in config.TEAMS[i].team_members else [person.ratings[i]][0]
                   for i in range(len(person.ratings))]
            row.insert(0, person.email)
            writer.writerow(row)

    df_new = pd.read_csv(
        "../../Fitness.csv")

    # Saving xlsx file. The xlsx file allows for the use of colored cells. In this case, we use
    # it to color the cell that corresponds to the team the person is assigned to.
    GFG = pd.ExcelWriter(
        globals.FITNESS_SCORES)
    df_new.to_excel(GFG, index=False)
    GFG.save()

    my_path = globals.FITNESS_SCORES
    document = load_workbook(my_path)
    sheet = document.active
    redFill = PatternFill(patternType='solid',
                          fgColor=colors.Color(rgb='00FF0000'))
    for i in range(2, sheet.max_row + 1):
        for j in range(2, sheet.max_column + 1):
            cell_value = list(str(sheet.cell(row=i, column=j).value))
            # If the cell is marked as the team this person is assigned to, fill the cell
            # with red.
            if "&" in cell_value:
                cell_value = cell_value[0:-1]
                sheet.cell(row=i, column=j).fill = redFill
            string = ""
            for element in cell_value:
                string += element
            c1 = sheet.cell(row=i, column=j)
            c1.value = int(string)

    document.save(
        filename=globals.FITNESS_SCORES)
    file = "../../Fitness.csv"
    # Delete the csv file that was just created
    if(os.path.exists(file) and os.path.isfile(file)):
        os.remove(file)
        print("file deleted")
    else:
        print("file not found")

    # This code creates the file that contains the names of the projects and the
    # people that were assigned to them
    with open(globals.TEAMS_FILE, "w", newline='') as f:
        writer = csv.writer(f)
        for team in config.TEAMS:
            row = [[member.email][0] for member in team.team_members]
            row.insert(0, team.identifier)
            row.insert(1, team.team_name)
            writer.writerow(row)

    # Store the emails and the project numbers the students were assigned to.
    with open(globals.STUDENTS, "w", newline='') as f:
        writer = csv.writer(f)
        people_copy = sorted(config.PEOPLE[:], key=lambda x: x.name)
        for person in people_copy:
            row = [config.TEAMS[i].identifier for i in range(config.LENGTH_PROJECT)
                   if person in config.TEAMS[i].team_members]
            row.insert(0, person.email)
            writer.writerow(row)

    with open(globals.TEAMS, "w", newline='') as f:
        writer = csv.writer(f)
        headers = []
        offset = 0
        # Adds all of the col names to the excel file
        for i in range(config.LENGTH_PROJECT):
            if i != 0 and config.PROJECT_NAMES[i - 1] == config.PROJECT_NAMES[i]:
                offset += 1
            headers.append("Project " + str(i + 1 - offset))
        writer.writerow(headers)
        for i in range(globals.MAX_TEAM_SIZE):
            row = [team.team_members[i].email if len(
                team) > i else None for team in config.TEAMS]
            writer.writerow(row)
